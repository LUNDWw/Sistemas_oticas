import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime

def generate_excel_report(orders):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Vendas"
    
    # Headers
    headers = ['OS', 'Cliente', 'Telefone', 'Loja', 'Laboratório', 'Data Exame', 'Status', 'Valor Total']
    ws.append(headers)
    
    # Style Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4361ee", end_color="4361ee", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    # Data
    for order in orders:
        total = (order['valor_pago'] or 0) + (order['entrada'] or 0) + (order['valor_retirada'] or 0)
        ws.append([
            order['os_number'],
            order['client_name'],
            order['phone'],
            order['store'],
            order['lab'],
            order['exam_date'],
            order['payment_status'],
            total
        ])
        
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_pdf_report(orders, title="Relatório de Vendas"):
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table Data
    data = [['OS', 'Cliente', 'Loja', 'Status', 'Valor']]
    total_sum = 0
    
    for order in orders:
        val = (order['valor_pago'] or 0) + (order['entrada'] or 0) + (order['valor_retirada'] or 0)
        total_sum += val
        data.append([
            str(order['os_number']),
            str(order['client_name'])[:20], # Truncate name
            str(order['store']),
            str(order['payment_status']),
            f"R$ {val:.2f}"
        ])
        
    # Style
    table = Table(data, colWidths=[60, 180, 80, 80, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4361ee')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"<b>Total Geral: R$ {total_sum:.2f}</b>", styles['Normal']))
    
    doc.build(elements)
    output.seek(0)
    return output
